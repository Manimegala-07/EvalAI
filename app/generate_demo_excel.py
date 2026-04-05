"""
Generate demo Excel file for EvalAI review/presentation
Run: python -m app.generate_demo_excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Data ──────────────────────────────────────────────────────
CO_OUTCOMES = {
    "CO1": "Knowledge & Recall — Recall facts, definitions, and basic concepts",
    "CO2": "Comprehension — Explain and describe concepts in own words",
    "CO3": "Application — Apply concepts to solve problems",
    "CO4": "Analysis — Break down, compare, and examine relationships",
    "CO5": "Evaluation — Justify, assess, and make judgments",
    "CO6": "Creation — Design, construct, or produce something new",
}

BLOOMS = {
    "L1": "Remember — Recall facts",
    "L2": "Understand — Explain concepts",
    "L3": "Apply — Use in new situations",
    "L4": "Analyze — Break down and compare",
    "L5": "Evaluate — Justify and assess",
    "L6": "Create — Design and produce",
}

DIFFICULTY = {
    1: "Very Easy",
    2: "Easy",
    3: "Average",
    4: "Hard",
    5: "Very Hard",
}

questions_data = [
    {
        "no": 1,
        "question": "What is a variable in Java?",
        "reference_answer": "A variable in Java is a named memory location that stores a value. It must be declared with a data type before use. Variables can store different types of data such as integers, strings, and decimals.",
        "difficulty": 1, "blooms": "L1", "co": "CO1",
        "students": [
            ("Arun Kumar",   "A variable in Java is a named memory location that stores a value. It must be declared with a data type before use. Variables can hold integers, strings, and other data types.",  "Fully Correct"),
            ("Priya Sharma", "A variable stores a value in memory and has a data type. We declare it before using it in the program.",                                                                           "Above Average"),
            ("Rahul Singh",  "Variable is used to store values in Java program.",                                                                                                                              "Average"),
            ("Divya Nair",   "A variable is a name given to store something.",                                                                                                                                 "Below Average"),
            ("Karthik Raja", "Variable is a function that returns values in Java.",                                                                                                                            "Fully Wrong"),
        ]
    },
    {
        "no": 2,
        "question": "What is the difference between int and double in Java?",
        "reference_answer": "int stores whole numbers without decimal points while double stores numbers with decimal points. int uses 4 bytes of memory and double uses 8 bytes. int is used for counting while double is used for precise calculations.",
        "difficulty": 2, "blooms": "L2", "co": "CO1",
        "students": [
            ("Arun Kumar",   "int stores whole numbers without decimals while double stores numbers with decimal points. int uses 4 bytes and double uses 8 bytes of memory.",                                  "Fully Correct"),
            ("Priya Sharma", "int is for whole numbers and double is for decimal numbers. double takes more memory than int.",                                                                                  "Above Average"),
            ("Rahul Singh",  "int and double are data types. int is for numbers and double is for bigger numbers.",                                                                                            "Average"),
            ("Divya Nair",   "int and double both store numbers in Java.",                                                                                                                                     "Below Average"),
            ("Karthik Raja", "int is used for text and double is used for numbers in Java.",                                                                                                                   "Fully Wrong"),
        ]
    },
    {
        "no": 3,
        "question": "Explain how a for loop works in Java.",
        "reference_answer": "A for loop in Java executes a block of code repeatedly for a fixed number of times. It has three parts: initialization which sets the starting value, condition which checks if the loop should continue, and increment which updates the counter. The loop stops when the condition becomes false.",
        "difficulty": 2, "blooms": "L2", "co": "CO2",
        "students": [
            ("Arun Kumar",   "A for loop executes a block of code repeatedly for a fixed number of times. It has three parts: initialization sets the starting value, condition checks if loop should continue, and increment updates the counter.",  "Fully Correct"),
            ("Priya Sharma", "For loop repeats code. It has init, condition and increment parts. It stops when condition becomes false.",                                                                       "Above Average"),
            ("Rahul Singh",  "For loop is used to repeat something in Java for a certain number of times.",                                                                                                    "Average"),
            ("Divya Nair",   "For loop repeats code many times in Java.",                                                                                                                                      "Below Average"),
            ("Karthik Raja", "For loop is used to declare variables in Java.",                                                                                                                                 "Fully Wrong"),
        ]
    },
    {
        "no": 4,
        "question": "How does exception handling work in Java?",
        "reference_answer": "Exception handling in Java uses try, catch, and finally blocks. Code that may throw an exception is placed in the try block. If an exception occurs, the catch block handles it and prevents the program from crashing. The finally block always executes regardless of whether an exception occurred.",
        "difficulty": 4, "blooms": "L3", "co": "CO3",
        "students": [
            ("Arun Kumar",   "Exception handling uses try, catch, and finally blocks. Code that may throw exception goes in try block. If exception occurs catch block handles it and prevents crash. Finally block always executes.",  "Fully Correct"),
            ("Priya Sharma", "Exception handling uses try and catch. Try block has risky code and catch handles the exception. Finally always runs.",                                                           "Above Average"),
            ("Rahul Singh",  "Exception handling is used to handle errors in Java using try catch blocks.",                                                                                                    "Average"),
            ("Divya Nair",   "Try catch is used for exceptions in Java.",                                                                                                                                      "Below Average"),
            ("Karthik Raja", "Exception handling stops the program from running when there is an error.",                                                                                                      "Fully Wrong"),
        ]
    },
    {
        "no": 5,
        "question": "What is the difference between ArrayList and LinkedList in Java?",
        "reference_answer": "ArrayList uses a dynamic array to store elements and provides fast random access using index. LinkedList uses a doubly linked list structure and provides fast insertion and deletion. ArrayList is better for frequent access operations while LinkedList is better for frequent insertions and deletions.",
        "difficulty": 4, "blooms": "L4", "co": "CO4",
        "students": [
            ("Arun Kumar",   "ArrayList uses dynamic array for fast random access using index. LinkedList uses doubly linked list for fast insertion and deletion. ArrayList is better for access and LinkedList for insertions.",  "Fully Correct"),
            ("Priya Sharma", "ArrayList is array based and LinkedList is node based. ArrayList is faster for access and LinkedList is faster for insert and delete.",                                           "Above Average"),
            ("Rahul Singh",  "ArrayList and LinkedList are both lists. ArrayList uses array and LinkedList uses links between nodes.",                                                                          "Average"),
            ("Divya Nair",   "ArrayList and LinkedList store data differently in Java.",                                                                                                                       "Below Average"),
            ("Karthik Raja", "ArrayList is faster than LinkedList for all operations in Java.",                                                                                                                "Fully Wrong"),
        ]
    },
    {
        "no": 6,
        "question": "How would you use an array to store and calculate the average of 5 student marks?",
        "reference_answer": "To store 5 student marks we declare an integer array of size 5. We then assign marks to each index from 0 to 4. To calculate average we use a for loop to sum all elements and divide by 5. This gives the average mark of all students.",
        "difficulty": 3, "blooms": "L3", "co": "CO3",
        "students": [
            ("Arun Kumar",   "We declare an integer array of size 5 to store marks. Assign marks to each index from 0 to 4. Use for loop to sum all elements and divide by 5 to get average mark.",            "Fully Correct"),
            ("Priya Sharma", "Create an array of size 5, store marks in it, loop through to add all marks and divide by 5 for average.",                                                                       "Above Average"),
            ("Rahul Singh",  "Use array to store marks and loop to calculate sum then divide by number of students.",                                                                                          "Average"),
            ("Divya Nair",   "Store marks in array and add them to get average.",                                                                                                                              "Below Average"),
            ("Karthik Raja", "Use a variable to store marks and calculate average by adding all.",                                                                                                             "Fully Wrong"),
        ]
    },
    {
        "no": 7,
        "question": "Compare the performance of bubble sort and merge sort algorithms.",
        "reference_answer": "Bubble sort has a time complexity of O(n²) in worst case making it inefficient for large datasets. Merge sort has a time complexity of O(n log n) making it much more efficient. Bubble sort is simple to implement but slow while merge sort is faster but requires extra memory space for merging.",
        "difficulty": 5, "blooms": "L4", "co": "CO4",
        "students": [
            ("Arun Kumar",   "Bubble sort has O(n²) time complexity making it inefficient for large data. Merge sort has O(n log n) which is much better. Bubble sort is simple but slow while merge sort is faster but needs extra memory.",  "Fully Correct"),
            ("Priya Sharma", "Bubble sort is slower with O(n²) and merge sort is faster with O(n log n). Merge sort needs more memory.",                                                                       "Above Average"),
            ("Rahul Singh",  "Bubble sort compares adjacent elements and merge sort divides and merges. Merge sort is faster.",                                                                                 "Average"),
            ("Divya Nair",   "Bubble sort is slow and merge sort is fast for sorting.",                                                                                                                        "Below Average"),
            ("Karthik Raja", "Bubble sort is better than merge sort because it is simpler to code.",                                                                                                           "Fully Wrong"),
        ]
    },
    {
        "no": 8,
        "question": "Evaluate whether using a static variable is a good practice in Java. Justify your answer.",
        "reference_answer": "Static variables are shared across all instances of a class which can be useful for constants or counters. However overusing static variables is bad practice because it creates tight coupling, makes testing difficult, and can cause unexpected behavior in multithreaded applications. Static variables should be used sparingly and only when truly needed.",
        "difficulty": 5, "blooms": "L5", "co": "CO5",
        "students": [
            ("Arun Kumar",   "Static variables are shared across all instances which is useful for constants and counters. But overusing them is bad practice because it creates tight coupling, makes testing hard, and causes issues in multithreaded apps.",  "Fully Correct"),
            ("Priya Sharma", "Static variables are useful for shared data but can cause problems in multithreading. They should be used carefully.",                                                            "Above Average"),
            ("Rahul Singh",  "Static variables are class level variables shared by all objects. They are useful sometimes but can cause problems.",                                                             "Average"),
            ("Divya Nair",   "Static variables are used when we need to share data between objects in Java.",                                                                                                  "Below Average"),
            ("Karthik Raja", "Static variables are always good to use because they save memory in Java.",                                                                                                      "Fully Wrong"),
        ]
    },
    {
        "no": 9,
        "question": "What is a string in Java and what makes it immutable?",
        "reference_answer": "A String in Java is a sequence of characters and is an object of the String class. Strings are immutable meaning once created their value cannot be changed. When you modify a string Java creates a new string object instead of changing the original. This immutability makes strings thread safe and allows string pooling for memory efficiency.",
        "difficulty": 3, "blooms": "L2", "co": "CO2",
        "students": [
            ("Arun Kumar",   "String in Java is a sequence of characters and object of String class. Strings are immutable meaning their value cannot be changed after creation. Modifying a string creates a new object. Immutability makes strings thread safe and allows string pooling.",  "Fully Correct"),
            ("Priya Sharma", "String is a sequence of characters. It is immutable so we cannot change it after creation. Java creates new string when we modify.",                                              "Above Average"),
            ("Rahul Singh",  "String stores text in Java. It is immutable which means it cannot be changed.",                                                                                                  "Average"),
            ("Divya Nair",   "String is used to store words and sentences in Java.",                                                                                                                           "Below Average"),
            ("Karthik Raja", "String is mutable in Java and can be changed anytime using methods.",                                                                                                            "Fully Wrong"),
        ]
    },
    {
        "no": 10,
        "question": "Design a simple class in Java to represent a Student with name, age, and grade attributes.",
        "reference_answer": "A Student class in Java should have private attributes name, age, and grade to follow encapsulation. We add a constructor to initialize these values and getter and setter methods to access and modify them. This design follows object oriented principles of encapsulation and abstraction making the code maintainable and reusable.",
        "difficulty": 5, "blooms": "L6", "co": "CO6",
        "students": [
            ("Arun Kumar",   "Student class should have private attributes name, age, and grade for encapsulation. Add constructor to initialize values and getter setter methods to access them. This follows OOP principles of encapsulation and abstraction.",  "Fully Correct"),
            ("Priya Sharma", "Create Student class with name, age, grade as private fields. Add constructor and getters setters. This follows encapsulation principle.",                                        "Above Average"),
            ("Rahul Singh",  "Student class has name, age, grade attributes. We use constructor to set values and methods to get them.",                                                                       "Average"),
            ("Divya Nair",   "Make a class Student with name age grade and add methods to get and set values.",                                                                                                "Below Average"),
            ("Karthik Raja", "Student class needs name and age variables and a method to print them.",                                                                                                         "Fully Wrong"),
        ]
    },
]

# ── Colors ─────────────────────────────────────────────────────
COLORS = {
    "header_blue":    "2D5BE3",
    "header_green":   "16A34A",
    "header_purple":  "7C3AED",
    "header_orange":  "D97706",
    "header_red":     "DC2626",
    "light_blue":     "EEF2FF",
    "light_green":    "DCFCE7",
    "light_yellow":   "FEF3C7",
    "light_red":      "FEE2E2",
    "light_purple":   "F3E8FF",
    "white":          "FFFFFF",
    "gray":           "F8F9FC",
    "dark":           "1A1A2E",
}

SCENARIO_COLORS = {
    "Fully Correct":  "DCFCE7",
    "Above Average":  "EEF2FF",
    "Average":        "FEF3C7",
    "Below Average":  "FED7AA",
    "Fully Wrong":    "FEE2E2",
}

def header_style(ws, cell, text, bg_color, font_color="FFFFFF", size=11, bold=True):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    ws[cell].fill = PatternFill("solid", fgColor=bg_color)
    ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def cell_style(ws, cell, text, bg_color=None, font_color="1A1A2E", bold=False, wrap=True, align="left"):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, color=font_color, size=10, name="Calibri")
    if bg_color:
        ws[cell].fill = PatternFill("solid", fgColor=bg_color)
    ws[cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def thin_border():
    thin = Side(style="thin", color="E8E6F0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — Question Bank
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Question Bank"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Title
ws1.merge_cells("A1:H1")
ws1["A1"] = "EvalAI — Question Bank"
ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws1["A1"].fill = PatternFill("solid", fgColor=COLORS["header_blue"])
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# Headers
headers = ["Q.No", "Question", "Reference Answer", "Difficulty", "Difficulty Label", "Bloom's Level", "Bloom's Description", "CO Mapping"]
cols    = ["A",    "B",        "C",                "D",          "E",                "F",             "G",                   "H"]
bg      = [COLORS["header_blue"]] * 8

for col, h, bg_c in zip(cols, headers, bg):
    header_style(ws1, f"{col}2", h, bg_c)
ws1.row_dimensions[2].height = 28

# Column widths
widths = [6, 45, 60, 10, 14, 12, 28, 12]
for col, w in zip(cols, widths):
    ws1.column_dimensions[col].width = w

# Data
for i, qd in enumerate(questions_data):
    row = i + 3
    bg_c = COLORS["gray"] if i % 2 == 0 else COLORS["white"]

    cell_style(ws1, f"A{row}", qd["no"],       bg_c, align="center", bold=True)
    cell_style(ws1, f"B{row}", qd["question"], bg_c, bold=True)
    cell_style(ws1, f"C{row}", qd["reference_answer"], bg_c)
    cell_style(ws1, f"D{row}", qd["difficulty"], bg_c, align="center", bold=True)
    cell_style(ws1, f"E{row}", DIFFICULTY[qd["difficulty"]], bg_c, align="center")
    cell_style(ws1, f"F{row}", qd["blooms"],   bg_c, align="center", bold=True)
    cell_style(ws1, f"G{row}", BLOOMS[qd["blooms"]], bg_c)
    cell_style(ws1, f"H{row}", qd["co"],       bg_c, align="center", bold=True)

    ws1.row_dimensions[row].height = 60

    for col in cols:
        ws1[f"{col}{row}"].border = thin_border()

# ══════════════════════════════════════════════════════════════
# SHEET 2 — Student Answers
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Student Answers")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:G1")
ws2["A1"] = "EvalAI — Student Answers (5 Scenarios per Question)"
ws2["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws2["A1"].fill = PatternFill("solid", fgColor=COLORS["header_green"])
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

headers2 = ["Q.No", "Question", "Student Name", "Scenario", "Student Answer", "Difficulty", "CO"]
cols2    = ["A",    "B",        "C",            "D",        "E",              "F",          "G"]

for col, h in zip(cols2, headers2):
    header_style(ws2, f"{col}2", h, COLORS["header_green"])
ws2.row_dimensions[2].height = 28

widths2 = [6, 40, 16, 14, 65, 10, 8]
for col, w in zip(cols2, widths2):
    ws2.column_dimensions[col].width = w

row = 3
for qd in questions_data:
    for student_name, answer, scenario in qd["students"]:
        bg_c = SCENARIO_COLORS.get(scenario, COLORS["white"])
        cell_style(ws2, f"A{row}", qd["no"],       bg_c, align="center", bold=True)
        cell_style(ws2, f"B{row}", qd["question"], bg_c)
        cell_style(ws2, f"C{row}", student_name,   bg_c, bold=True)
        cell_style(ws2, f"D{row}", scenario,        bg_c, bold=True, align="center")
        cell_style(ws2, f"E{row}", answer,          bg_c)
        cell_style(ws2, f"F{row}", f"D{qd['difficulty']} — {DIFFICULTY[qd['difficulty']]}", bg_c, align="center")
        cell_style(ws2, f"G{row}", qd["co"],        bg_c, align="center", bold=True)
        ws2.row_dimensions[row].height = 55
        for col in cols2:
            ws2[f"{col}{row}"].border = thin_border()
        row += 1

# ══════════════════════════════════════════════════════════════
# SHEET 3 — CO Outcomes
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("CO Outcomes")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
ws3["A1"] = "EvalAI — Course Outcomes (CO) Definitions"
ws3["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws3["A1"].fill = PatternFill("solid", fgColor=COLORS["header_purple"])
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

for col, h, bg_c in zip(["A","B","C"], ["CO", "Title", "Description"], [COLORS["header_purple"]]*3):
    header_style(ws3, f"{col}2", h, bg_c)
ws3.row_dimensions[2].height = 28
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 60

co_colors = ["EEF2FF","F3E8FF","DCFCE7","FEF3C7","FED7AA","FEE2E2"]
for i, (co, desc) in enumerate(CO_OUTCOMES.items()):
    row = i + 3
    parts = desc.split(" — ")
    bg_c = co_colors[i]
    cell_style(ws3, f"A{row}", co,       bg_c, bold=True, align="center")
    cell_style(ws3, f"B{row}", parts[0], bg_c, bold=True)
    cell_style(ws3, f"C{row}", parts[1] if len(parts) > 1 else "", bg_c)
    ws3.row_dimensions[row].height = 36
    for col in ["A","B","C"]:
        ws3[f"{col}{row}"].border = thin_border()

# ══════════════════════════════════════════════════════════════
# SHEET 4 — Bloom's Taxonomy
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Blooms Taxonomy")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:D1")
ws4["A1"] = "EvalAI — Bloom's Taxonomy Levels"
ws4["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws4["A1"].fill = PatternFill("solid", fgColor=COLORS["header_orange"])
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

for col, h in zip(["A","B","C","D"], ["Level", "Name", "Description", "Question Stem Examples"]):
    header_style(ws4, f"{col}2", h, COLORS["header_orange"])
ws4.row_dimensions[2].height = 28
ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 45

blooms_full = [
    ("L1", "Remember",   "Recall facts and basic concepts",              "What is...? Define... List... Name... State..."),
    ("L2", "Understand", "Explain ideas or concepts in own words",       "Explain... Describe... Summarize... How does...?"),
    ("L3", "Apply",      "Use information in new situations",            "How would you use...? Apply... Calculate... Solve..."),
    ("L4", "Analyze",    "Draw connections and break down information",  "Compare... Analyze... Differentiate... Break down..."),
    ("L5", "Evaluate",   "Justify a decision or course of action",       "Evaluate... Justify... Assess... Why is... better?"),
    ("L6", "Create",     "Produce new or original work",                 "Design... Create... Develop... Construct... Build..."),
]

bloom_colors = ["F0FDF4","EEF2FF","FEF3C7","F3E8FF","FED7AA","FEE2E2"]
for i, (level, name, desc, stems) in enumerate(blooms_full):
    row = i + 3
    bg_c = bloom_colors[i]
    cell_style(ws4, f"A{row}", level, bg_c, bold=True, align="center")
    cell_style(ws4, f"B{row}", name,  bg_c, bold=True)
    cell_style(ws4, f"C{row}", desc,  bg_c)
    cell_style(ws4, f"D{row}", stems, bg_c)
    ws4.row_dimensions[row].height = 40
    for col in ["A","B","C","D"]:
        ws4[f"{col}{row}"].border = thin_border()

# ══════════════════════════════════════════════════════════════
# SHEET 5 — Summary
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:E1")
ws5["A1"] = "EvalAI — Demo Data Summary"
ws5["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws5["A1"].fill = PatternFill("solid", fgColor=COLORS["dark"])
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 36

for col, h in zip(["A","B","C","D","E"], ["Q.No", "Question (Short)", "Difficulty", "Bloom's", "CO"]):
    header_style(ws5, f"{col}2", h, COLORS["dark"])
ws5.row_dimensions[2].height = 28
ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 50
ws5.column_dimensions["C"].width = 18
ws5.column_dimensions["D"].width = 12
ws5.column_dimensions["E"].width = 8

diff_colors_map = {1:"DCFCE7", 2:"D1FAE5", 3:"FEF3C7", 4:"FED7AA", 5:"FEE2E2"}
for i, qd in enumerate(questions_data):
    row = i + 3
    bg_c = diff_colors_map[qd["difficulty"]]
    cell_style(ws5, f"A{row}", qd["no"],                                    bg_c, align="center", bold=True)
    cell_style(ws5, f"B{row}", qd["question"],                              bg_c)
    cell_style(ws5, f"C{row}", f"D{qd['difficulty']} — {DIFFICULTY[qd['difficulty']]}", bg_c, align="center")
    cell_style(ws5, f"D{row}", qd["blooms"],                                bg_c, align="center", bold=True)
    cell_style(ws5, f"E{row}", qd["co"],                                    bg_c, align="center", bold=True)
    ws5.row_dimensions[row].height = 40
    for col in ["A","B","C","D","E"]:
        ws5[f"{col}{row}"].border = thin_border()

# ── Save ───────────────────────────────────────────────────────
output_path = "data/EvalAI_Demo_Data.xlsx"
import os
os.makedirs("data", exist_ok=True)
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print(f"Sheets: Question Bank | Student Answers | CO Outcomes | Blooms Taxonomy | Summary")
