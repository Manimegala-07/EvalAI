"""
Generate Tamil + English demo Excel file for EvalAI review/presentation
Run: python -m app.generate_tamil_excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

DIFFICULTY = {1: "Very Easy", 2: "Easy", 3: "Average", 4: "Hard", 5: "Very Hard"}
BLOOMS = {
    "L1": "Remember", "L2": "Understand", "L3": "Apply",
    "L4": "Analyze",  "L5": "Evaluate",  "L6": "Create",
}

questions_data = [
    {
        "no": 1, "difficulty": 1, "blooms": "L1", "co": "CO1",
        "question_en": "What is a variable in Java?",
        "question_ta": "ஜாவாவில் மாறி என்றால் என்ன?",
        "ref_en": "A variable in Java is a named memory location that stores a value. It must be declared with a data type before use.",
        "ref_ta": "ஜாவாவில் மாறி என்பது ஒரு மதிப்பை சேமிக்கும் பெயரிடப்பட்ட நினைவக இடம். பயன்படுத்துவதற்கு முன் தரவு வகையுடன் அறிவிக்கப்பட வேண்டும்.",
        "students": [
            ("Fully Correct",  "ஜாவாவில் மாறி என்பது ஒரு மதிப்பை சேமிக்கும் பெயரிடப்பட்ட நினைவக இடம். பயன்படுத்துவதற்கு முன் தரவு வகையுடன் அறிவிக்கப்பட வேண்டும். முழு எண்கள், சரங்கள் போன்ற தரவுகளை சேமிக்கலாம்.", "9-10"),
            ("Above Average",  "மாறி என்பது நினைவகத்தில் மதிப்பை சேமிக்கும் இடம். தரவு வகையுடன் அறிவிக்க வேண்டும்.", "7-8"),
            ("Average",        "மாறி என்பது ஜாவாவில் மதிப்புகளை சேமிக்க பயன்படுகிறது.", "5-6"),
            ("Below Average",  "மாறி ஒரு பெயர்.", "3-4"),
            ("Fully Wrong",    "மாறி என்பது ஒரு செயல்பாடு ஆகும்.", "1-2"),
        ]
    },
    {
        "no": 2, "difficulty": 2, "blooms": "L2", "co": "CO1",
        "question_en": "Explain the difference between int and double in Java.",
        "question_ta": "ஜாவாவில் int மற்றும் double இடையே உள்ள வேறுபாட்டை விளக்குக.",
        "ref_en": "int stores whole numbers without decimal points while double stores numbers with decimal points. int uses 4 bytes and double uses 8 bytes of memory.",
        "ref_ta": "int என்பது தசம புள்ளிகள் இல்லாத முழு எண்களை சேமிக்கிறது, double என்பது தசம புள்ளிகளுடன் எண்களை சேமிக்கிறது. int 4 bytes மற்றும் double 8 bytes நினைவகம் பயன்படுத்துகிறது.",
        "students": [
            ("Fully Correct",  "int தசம புள்ளிகள் இல்லாத முழு எண்களை சேமிக்கிறது, double தசம புள்ளிகளுடன் எண்களை சேமிக்கிறது. int 4 bytes மற்றும் double 8 bytes நினைவகம் பயன்படுத்துகிறது.", "8-9"),
            ("Above Average",  "int முழு எண்களுக்கும் double தசம எண்களுக்கும் பயன்படுகிறது. double அதிக நினைவகம் எடுக்கும்.", "6-7"),
            ("Average",        "int மற்றும் double இரண்டும் எண்களை சேமிக்கும் தரவு வகைகள்.", "4-5"),
            ("Below Average",  "int மற்றும் double எண்களுக்கு பயன்படுகிறது.", "2-3"),
            ("Fully Wrong",    "int உரைக்கும் double எண்களுக்கும் பயன்படுகிறது.", "1"),
        ]
    },
    {
        "no": 3, "difficulty": 3, "blooms": "L3", "co": "CO3",
        "question_en": "How would you use a for loop to print numbers 1 to 10 in Java?",
        "question_ta": "ஜாவாவில் 1 முதல் 10 வரை எண்களை அச்சிட for loop எவ்வாறு பயன்படுத்துவீர்கள்?",
        "ref_en": "We use a for loop with initialization i=1, condition i<=10, and increment i++. Inside the loop we use System.out.println to print each value of i. The loop runs 10 times printing numbers 1 to 10.",
        "ref_ta": "i=1 என துவக்கம், i<=10 என நிபந்தனை மற்றும் i++ என அதிகரிப்புடன் for loop பயன்படுத்துகிறோம். loop உள்ளே System.out.println மூலம் i இன் ஒவ்வொரு மதிப்பையும் அச்சிடுகிறோம். loop 10 முறை இயங்கி 1 முதல் 10 வரை அச்சிடும்.",
        "students": [
            ("Fully Correct",  "i=1 என துவக்கம், i<=10 என நிபந்தனை, i++ என அதிகரிப்புடன் for loop பயன்படுத்துகிறோம். System.out.println மூலம் i ஐ அச்சிடுகிறோம். loop 10 முறை இயங்கும்.", "7-8"),
            ("Above Average",  "for loop இல் i=1 இலிருந்து i<=10 வரை i++ அதிகரிப்புடன் loop இட்டு println மூலம் அச்சிடலாம்.", "5-6"),
            ("Average",        "for loop பயன்படுத்தி 1 முதல் 10 வரை அச்சிடலாம்.", "4-5"),
            ("Below Average",  "loop பயன்படுத்தி எண்களை அச்சிடலாம்.", "2-3"),
            ("Fully Wrong",    "while loop பயன்படுத்தி மட்டுமே எண்களை அச்சிட முடியும்.", "1"),
        ]
    },
    {
        "no": 4, "difficulty": 4, "blooms": "L4", "co": "CO4",
        "question_en": "Compare ArrayList and LinkedList in Java with respect to performance.",
        "question_ta": "செயல்திறன் அடிப்படையில் ஜாவாவில் ArrayList மற்றும் LinkedList ஐ ஒப்பிடுக.",
        "ref_en": "ArrayList uses dynamic array providing fast random access with O(1) time. LinkedList uses doubly linked list providing fast insertion and deletion. ArrayList is slow for insertion while LinkedList is slow for random access.",
        "ref_ta": "ArrayList dynamic array பயன்படுத்தி O(1) நேரத்தில் வேகமான random access வழங்குகிறது. LinkedList doubly linked list பயன்படுத்தி வேகமான செருகல் மற்றும் நீக்கம் வழங்குகிறது. ArrayList செருகலில் மெதுவாகவும் LinkedList access இல் மெதுவாகவும் இருக்கும்.",
        "students": [
            ("Fully Correct",  "ArrayList dynamic array பயன்படுத்தி O(1) random access வழங்குகிறது. LinkedList doubly linked list பயன்படுத்தி வேகமான செருகல் நீக்கம் வழங்குகிறது. ArrayList செருகலில் மெதுவாகவும் LinkedList access இல் மெதுவாகவும் இருக்கும்.", "6-7"),
            ("Above Average",  "ArrayList random access க்கு வேகமானது. LinkedList செருகல் நீக்கத்திற்கு வேகமானது.", "4-5"),
            ("Average",        "ArrayList மற்றும் LinkedList இரண்டும் தரவை சேமிக்கும். ArrayList array அடிப்படையிலும் LinkedList node அடிப்படையிலும் செயல்படும்.", "3-4"),
            ("Below Average",  "ArrayList மற்றும் LinkedList வேறுபட்ட முறையில் தரவை சேமிக்கும்.", "2-3"),
            ("Fully Wrong",    "ArrayList எல்லா செயல்பாடுகளிலும் LinkedList ஐ விட வேகமானது.", "1"),
        ]
    },
    {
        "no": 5, "difficulty": 5, "blooms": "L5", "co": "CO5",
        "question_en": "Evaluate whether using static variables is good practice in Java. Justify your answer.",
        "question_ta": "ஜாவாவில் static மாறிகளை பயன்படுத்துவது நல்ல நடைமுறையா என மதிப்பீடு செய்து நியாயப்படுத்துக.",
        "ref_en": "Static variables are useful for constants and shared counters but overusing them creates tight coupling and makes testing difficult. In multithreaded applications static variables can cause unexpected behavior. They should be used sparingly only when truly necessary.",
        "ref_ta": "Static மாறிகள் மாறிலிகள் மற்றும் பகிரப்பட்ட கவுண்டர்களுக்கு பயனுள்ளவை ஆனால் அதிகமாக பயன்படுத்துவது tight coupling உருவாக்கி சோதனையை கடினமாக்கும். பல நூல் பயன்பாடுகளில் static மாறிகள் எதிர்பாராத நடத்தையை ஏற்படுத்தலாம். உண்மையிலேயே தேவைப்படும்போது மட்டுமே பயன்படுத்த வேண்டும்.",
        "students": [
            ("Fully Correct",  "Static மாறிகள் மாறிலிகளுக்கு பயனுள்ளவை ஆனால் அதிகமாக பயன்படுத்துவது tight coupling உருவாக்கும். பல நூல் பயன்பாடுகளில் எதிர்பாராத நடத்தை ஏற்படலாம். எனவே தேவைப்படும்போது மட்டுமே பயன்படுத்த வேண்டும்.", "5-6"),
            ("Above Average",  "Static மாறிகள் பகிரப்பட்ட தரவுக்கு பயனுள்ளவை ஆனால் multithreading இல் சிக்கல் ஏற்படலாம். கவனமாக பயன்படுத்த வேண்டும்.", "3-4"),
            ("Average",        "Static மாறிகள் class அளவில் பகிரப்படும். சில நேரங்களில் பயனுள்ளவை சில நேரங்களில் சிக்கல் ஏற்படுத்தும்.", "2-3"),
            ("Below Average",  "Static மாறிகள் objects இடையே தரவை பகிர பயன்படுகிறது.", "1-2"),
            ("Fully Wrong",    "Static மாறிகள் எப்போதும் நல்லவை ஏனெனில் நினைவகத்தை சேமிக்கும்.", "1"),
        ]
    },
]

SCENARIO_COLORS = {
    "Fully Correct":  "DCFCE7",
    "Above Average":  "EEF2FF",
    "Average":        "FEF3C7",
    "Below Average":  "FED7AA",
    "Fully Wrong":    "FEE2E2",
}

DIFF_COLORS = {1: "DCFCE7", 2: "D1FAE5", 3: "FEF3C7", 4: "FED7AA", 5: "FEE2E2"}

def hdr(ws, cell, text, bg, fg="FFFFFF", size=11, bold=True):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, color=fg, size=size, name="Calibri")
    ws[cell].fill = PatternFill("solid", fgColor=bg)
    ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def cel(ws, cell, text, bg=None, fg="1A1A2E", bold=False, align="left"):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, color=fg, size=10, name="Calibri")
    if bg:
        ws[cell].fill = PatternFill("solid", fgColor=bg)
    ws[cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def border():
    s = Side(style="thin", color="E8E6F0")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()

# ══ SHEET 1 — Question Bank (EN + TA) ══════════════════════════
ws1 = wb.active
ws1.title = "Question Bank"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

ws1.merge_cells("A1:I1")
ws1["A1"] = "EvalAI — Tamil + English Question Bank"
ws1["A1"].font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
ws1["A1"].fill = PatternFill("solid", fgColor="2D5BE3")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

cols1 = ["A","B","C","D","E","F","G","H","I"]
hdrs1 = ["Q.No","Question (English)","Question (Tamil)","Reference (English)","Reference (Tamil)","Difficulty","Level","Bloom's","CO"]
for c, h in zip(cols1, hdrs1):
    hdr(ws1, f"{c}2", h, "2D5BE3")
ws1.row_dimensions[2].height = 28

for c, w in zip(cols1, [5, 35, 35, 45, 45, 10, 12, 10, 8]):
    ws1.column_dimensions[c].width = w

for i, qd in enumerate(questions_data):
    row = i + 3
    bg = "F8F9FC" if i % 2 == 0 else "FFFFFF"
    cel(ws1, f"A{row}", qd["no"],           bg, align="center", bold=True)
    cel(ws1, f"B{row}", qd["question_en"],  bg, bold=True)
    cel(ws1, f"C{row}", qd["question_ta"],  bg, bold=True)
    cel(ws1, f"D{row}", qd["ref_en"],       bg)
    cel(ws1, f"E{row}", qd["ref_ta"],       bg)
    cel(ws1, f"F{row}", qd["difficulty"],   DIFF_COLORS[qd["difficulty"]], align="center", bold=True)
    cel(ws1, f"G{row}", DIFFICULTY[qd["difficulty"]], DIFF_COLORS[qd["difficulty"]], align="center")
    cel(ws1, f"H{row}", qd["blooms"],       bg, align="center", bold=True)
    cel(ws1, f"I{row}", qd["co"],           bg, align="center", bold=True)
    ws1.row_dimensions[row].height = 80
    for c in cols1:
        ws1[f"{c}{row}"].border = border()

# ══ SHEET 2 — Student Answers (Tamil) ══════════════════════════
ws2 = wb.create_sheet("Tamil Student Answers")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:G1")
ws2["A1"] = "EvalAI — Tamil Student Answers (5 Quality Scenarios)"
ws2["A1"].font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
ws2["A1"].fill = PatternFill("solid", fgColor="16A34A")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

cols2 = ["A","B","C","D","E","F","G"]
hdrs2 = ["Q.No","Question (Tamil)","Scenario","Student Answer (Tamil)","Difficulty","Bloom's","Expected Score"]
for c, h in zip(cols2, hdrs2):
    hdr(ws2, f"{c}2", h, "16A34A")
ws2.row_dimensions[2].height = 28

for c, w in zip(cols2, [5, 35, 14, 60, 14, 10, 14]):
    ws2.column_dimensions[c].width = w

row = 3
for qd in questions_data:
    for scenario, answer, score in qd["students"]:
        bg = SCENARIO_COLORS.get(scenario, "FFFFFF")
        cel(ws2, f"A{row}", qd["no"],          bg, align="center", bold=True)
        cel(ws2, f"B{row}", qd["question_ta"], bg)
        cel(ws2, f"C{row}", scenario,           bg, bold=True, align="center")
        cel(ws2, f"D{row}", answer,             bg)
        cel(ws2, f"E{row}", f"D{qd['difficulty']} — {DIFFICULTY[qd['difficulty']]}", DIFF_COLORS[qd["difficulty"]], align="center")
        cel(ws2, f"F{row}", qd["blooms"],       bg, align="center", bold=True)
        cel(ws2, f"G{row}", score,              bg, align="center", bold=True)
        ws2.row_dimensions[row].height = 60
        for c in cols2:
            ws2[f"{c}{row}"].border = border()
        row += 1

# ══ SHEET 3 — Score Pattern ═════════════════════════════════════
ws3 = wb.create_sheet("Score Pattern")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
ws3["A1"] = "EvalAI — Expected Score Pattern by Difficulty"
ws3["A1"].font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
ws3["A1"].fill = PatternFill("solid", fgColor="7C3AED")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

cols3 = ["A","B","C","D","E","F","G"]
hdrs3 = ["Difficulty","Level","Fully Correct","Above Average","Average","Below Average","Fully Wrong"]
for c, h in zip(cols3, hdrs3):
    hdr(ws3, f"{c}2", h, "7C3AED")
ws3.row_dimensions[2].height = 28

for c, w in zip(cols3, [10, 12, 14, 14, 12, 14, 12]):
    ws3.column_dimensions[c].width = w

score_pattern = [
    (1, "Very Easy",  "9-10", "7-8", "5-6", "3-4", "1-2"),
    (2, "Easy",       "8-9",  "6-7", "4-5", "2-3", "1"),
    (3, "Average",    "7-8",  "5-6", "4-5", "2-3", "1"),
    (4, "Hard",       "6-7",  "4-5", "3-4", "2-3", "1"),
    (5, "Very Hard",  "5-6",  "3-4", "2-3", "1-2", "1"),
]

for i, (d, label, fc, aa, avg, ba, fw) in enumerate(score_pattern):
    row = i + 3
    bg = DIFF_COLORS[d]
    cel(ws3, f"A{row}", f"D{d}", bg, align="center", bold=True)
    cel(ws3, f"B{row}", label,   bg, align="center")
    cel(ws3, f"C{row}", fc,      "DCFCE7", align="center", bold=True)
    cel(ws3, f"D{row}", aa,      "EEF2FF", align="center", bold=True)
    cel(ws3, f"E{row}", avg,     "FEF3C7", align="center", bold=True)
    cel(ws3, f"F{row}", ba,      "FED7AA", align="center", bold=True)
    cel(ws3, f"G{row}", fw,      "FEE2E2", align="center", bold=True)
    ws3.row_dimensions[row].height = 32
    for c in cols3:
        ws3[f"{c}{row}"].border = border()

# ══ SHEET 4 — Summary ═══════════════════════════════════════════
ws4 = wb.create_sheet("Summary")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
ws4["A1"] = "EvalAI — Tamil Question Set Summary"
ws4["A1"].font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
ws4["A1"].fill = PatternFill("solid", fgColor="1A1A2E")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

cols4 = ["A","B","C","D","E","F"]
hdrs4 = ["Q.No","Question (English)","Difficulty","Bloom's","CO","Score Range (Fully Correct)"]
for c, h in zip(cols4, hdrs4):
    hdr(ws4, f"{c}2", h, "1A1A2E")
ws4.row_dimensions[2].height = 28

for c, w in zip(cols4, [5, 50, 16, 10, 8, 22]):
    ws4.column_dimensions[c].width = w

score_ranges = ["9-10", "8-9", "7-8", "6-7", "5-6"]
for i, qd in enumerate(questions_data):
    row = i + 3
    bg = DIFF_COLORS[qd["difficulty"]]
    cel(ws4, f"A{row}", qd["no"],                                       bg, align="center", bold=True)
    cel(ws4, f"B{row}", qd["question_en"],                              bg)
    cel(ws4, f"C{row}", f"D{qd['difficulty']} — {DIFFICULTY[qd['difficulty']]}", bg, align="center")
    cel(ws4, f"D{row}", qd["blooms"],                                   bg, align="center", bold=True)
    cel(ws4, f"E{row}", qd["co"],                                       bg, align="center", bold=True)
    cel(ws4, f"F{row}", score_ranges[i],                                "DCFCE7", align="center", bold=True)
    ws4.row_dimensions[row].height = 40
    for c in cols4:
        ws4[f"{c}{row}"].border = border()

os.makedirs("data", exist_ok=True)
wb.save("data/EvalAI_Tamil_Demo.xlsx")
print("Saved: data/EvalAI_Tamil_Demo.xlsx")
print("Sheets: Question Bank | Tamil Student Answers | Score Pattern | Summary")
